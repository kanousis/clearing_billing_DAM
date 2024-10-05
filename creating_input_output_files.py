# in this script we create the input and output files for the optimization problem
# we want to split the data of the original file into 24 different files, each file
# will contain the data of each offer for a specific hour of the day
import pandas as pd
from openpyxl import load_workbook
input_file = '20240910_EL-DAM_AggrCurves_EN_v01.xlsx'
for hour in range(1, 25):
    input_file_new = 'input_files/input_%s.xlsx' %hour
    output_file = 'output_files/output_%s.xlsx' %hour

    df = pd.read_excel(input_file, usecols="B,E,H,I")
    quantity_offers = []
    unitprice_offers = []
    quantity_bids = []
    unitprice_bids = []

    for i in range(len(df)):
        if df.iloc[i, 0] == "Sell" and df.iloc[i, 1] == hour:
            quantity_offers.append(df.iloc[i, 2])
            unitprice_offers.append(df.iloc[i, 3])

    for i in range(len(df)):
        if df.iloc[i, 0] == "Buy" and df.iloc[i, 1] == hour:
            quantity_bids.append(df.iloc[i, 2])
            unitprice_bids.append(df.iloc[i, 3])

    wb1 = load_workbook(input_file_new)
    ws1 = wb1.active

    # Clear previous data in the worksheet
    for row in ws1.iter_rows(min_row=3, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.value = None
    # Write new data to the worksheet
    for i in range (len(unitprice_offers)):
        ws1['B%s'%(i+3)].value = unitprice_offers[i]
    for i in range (len(quantity_offers)):
        ws1['C%s'%(i+3)].value = quantity_offers[i]
    for i in range (len(unitprice_bids)):
        ws1['F%s'%(i+3)].value = unitprice_bids[i]
    for i in range (len(quantity_bids)):
        ws1['G%s'%(i+3)].value = quantity_bids[i] 

    for i in range (len(quantity_offers)):
        ws1['A%s'%(i+3)].value = '-' 
    for i in range (len(quantity_bids)):
        ws1['E%s'%(i+3)].value = '-'

    wb1.save(input_file_new)
    wb2 = load_workbook(output_file)
    ws2 = wb2.active

    for i in range (len(quantity_offers)):
        ws2['B%s'%(i+3)].value = quantity_offers[i]
    for i in range (len(quantity_bids)):
        ws2['H%s'%(i+3)].value = quantity_bids[i]        
    for i in range (len(quantity_offers)):
        ws2['A%s'%(i+3)].value = '-' 
    for i in range (len(quantity_bids)):
        ws2['G%s'%(i+3)].value = '-'

    wb2.save(output_file)


    