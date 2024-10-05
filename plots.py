import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

def plot_total_energy_exchenged(total_energy_exchanged, hours):
    plt.figure(1)
    plt.plot(hours, total_energy_exchanged, marker='o', linestyle='-')
    plt.xticks(hours)
    plt.xlabel('Hour')
    plt.ylabel('Total Energy Exchanged(MWh)')
    plt.title('Total Energy Exchanged per Hour')
    

def plot_market_prices(market_prices, hours):
    plt.figure(2)
    plt.plot(hours, market_prices, marker='o', linestyle='-')
    plt.xticks(hours)
    plt.xlabel('Hour')
    plt.ylabel('Market Price(€/MWh)')
    plt.title('Market Price per Hour')
    

def plot_total_producers_profit(total_producers_profit, hours):
    plt.figure(3)
    plt.plot(hours, total_producers_profit, marker='o', linestyle='-')
    plt.xticks(hours)
    plt.xlabel('Hour')
    plt.ylabel('Total Producers Profit(€)')
    plt.title('Total Producers Profit per Hour')

def plot_total_buyers_profit(total_buyers_profit, hours):
    plt.figure(4)
    plt.plot(hours, total_buyers_profit, marker='o', linestyle='-')
    plt.xticks(hours)
    plt.xlabel('Hour')
    plt.ylabel('Total Buyers Profit(€)')
    plt.title('Total Buyers Profit per Hour')

def plot_social_welfare(social_welfare, hours):
    plt.figure(5)
    plt.plot(hours, social_welfare, marker='o', linestyle='-')
    plt.xticks(hours)
    plt.xlabel('Hour')
    plt.ylabel('Social Welfare(€)')
    plt.title('Social Welfare per Hour')
    plt.show()

def plot_welfare_profits(total_producers_profit, total_buyers_profit, social_welfare, hours):
    plt.figure(6)
    plt.plot(hours, total_producers_profit, marker='o', linestyle='-', label='Total Producers Profit')
    plt.plot(hours, total_buyers_profit, marker='o', linestyle='-', label='Total Buyers Profit')
    plt.plot(hours, social_welfare, marker='o', linestyle='-', label='Social Welfare')
    plt.xticks(hours)
    plt.xlabel('Hour')
    plt.ylabel('Value(€)')
    plt.title('Total Producers Profit, Total Buyers Profit and Social Welfare per Hour')
    plt.legend()
    plt.show()

hours = np.linspace(1, 24, 24)

total_energy_exchanged = []
# Load the output files and get the total energy exchanged
for hour in range(1, 25):
    output_file = 'output_files/output_%s.xlsx' %hour

    wb = load_workbook(output_file)
    ws = wb.active

    total_energy_exchanged.append(ws['N2'].value)

# Load total number of fully/partially/rejected offers and bids
fully_accepted_offers = []
partially_accepted_offers = []
rejected_offers = []

fully_accepted_bids = []
partially_accepted_bids = []
rejected_bids = []

for hour in range(1, 25):
    output_file = 'output_files/output_%s.xlsx' %hour

    wb = load_workbook(output_file)
    ws = wb.active

    fully_accepted_offers.append(ws['N8'].value)
    partially_accepted_offers.append(ws['N9'].value)
    rejected_offers.append(ws['N10'].value)

    fully_accepted_bids.append(ws['N12'].value)
    partially_accepted_bids.append(ws['N13'].value)
    rejected_bids.append(ws['N14'].value)

print('Total number of fully accepted offers:', fully_accepted_offers)
print('Total number of partially accepted offers:', partially_accepted_offers)
print('Total number of rejected offers:', rejected_offers)

print('Total number of fully accepted bids:', fully_accepted_bids)
print('Total number of partially accepted bids:', partially_accepted_bids)
print('Total number of rejected bids:', rejected_bids)


# Load marget price for every hour
market_prices = []
for hour in range(1, 25):
    output_file = 'output_files/output_%s.xlsx' %hour

    wb = load_workbook(output_file)
    ws = wb.active

    market_prices.append(ws['N3'].value)

# Load total producers profit/ total buyers profit/ social welfare, for every hour
total_producers_profit = []
total_buyers_profit = []
social_welfare = []

for hour in range(1, 25):
    output_file = 'output_files/output_%s.xlsx' %hour

    wb = load_workbook(output_file)
    ws = wb.active

    total_producers_profit.append(ws['N4'].value)
    total_buyers_profit.append(ws['N5'].value)
    social_welfare.append(ws['N6'].value)


plot_total_energy_exchenged(total_energy_exchanged, hours)
plot_market_prices(market_prices, hours)
plot_total_producers_profit(total_producers_profit, hours)
plot_total_buyers_profit(total_buyers_profit, hours)
plot_social_welfare(social_welfare, hours)
plot_welfare_profits(total_producers_profit, total_buyers_profit, social_welfare, hours)



