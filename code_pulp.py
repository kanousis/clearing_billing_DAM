import pandas as pd
import numpy as np
from pulp import LpMaximize, LpProblem, LpVariable, lpSum, LpStatus, value
from openpyxl import load_workbook
import time

t = time.time()
# Define the solver function using PuLP
def solver(c, Aub, bub, Aeq, beq):
    # Initialize the problem
    model = LpProblem("clearance_billing", LpMaximize)

    # Define variables
    n = len(c)
    x = [LpVariable(f'x_{i}', lowBound=0) for i in range(n)]

    # Objective function
    model += lpSum(c[i] * x[i] for i in range(n)), "Objective"

    # Inequality constraints
    for i in range(len(Aub)):
        model += lpSum(Aub[i][j] * x[j] for j in range(n)) <= bub[i], f"Inequality_{i}"

    # Equality constraints
    for i in range(len(Aeq)):
        model += lpSum(Aeq[i][j] * x[j] for j in range(n)) == beq[i], f"Equality_{i}"

    # Solve the problem
    model.solve()

    # Print the status of the solution
    print(f"Status: {LpStatus[model.status]}")

    # Calculate the value of the objective function (Social Welfare)
    SW = round(value(model.objective), 3)

    # Append the values of the decision variables to a list
    x_list_array = []
    for j in range(n):
        x_list_array.append(value(x[j]))
        x_list = np.array(x_list_array)

    # Calculate the value of the dual variable (shadow price), of the equality constraint,
    # which is the market price
    λ = round(- model.constraints[list(model.constraints.keys())[-1]].pi, 2)
    return SW, x_list, λ

for hour in range(1, 25):
    print("Hour:", hour)
    # File path to the Excel input file of each hour
    input_file = 'input_files/input_%s.xlsx' %hour
    
    # Read the Excel file
    sale_offers = pd.read_excel(input_file, index_col=0, usecols="A:C")
    purch_bids = pd.read_excel(input_file, index_col=0, usecols="E:G")

    # Extract from the Excel file the selling values
    offers = ((sale_offers.iloc[1:, 0]).dropna()).astype(float).values
    # Extract from the Excel file the bidding values 
    bids = ((purch_bids.iloc[1:, 0]).dropna()).astype(float).values

    # Extract from the Excel file the maximum power values
    offers_Power = ((sale_offers.iloc[1:, 1]).dropna()).astype(float).values
    bids_Power = ((purch_bids.iloc[1:, 1]).dropna()).astype(float).values

    # Set the parameters of the problem
    Cg = np.concatenate((offers, np.zeros(len(bids))))
    Bd = np.concatenate((np.zeros(len(offers)), bids))

    # Set the objective function coefficients
    c = Bd - Cg

    # Set the constraints
    Aub = np.eye(len(c))
    bub = np.concatenate((offers_Power, bids_Power))

    Aeq = np.concatenate((np.ones(len(bids)), -np.ones(len(offers)))).reshape(1, -1)
    beq = np.array([0])        


    # Call the solver function
    SW, x_list, λ = solver(c, Aub, bub, Aeq, beq)

    # calculate the total power sold/bougth 
    Total_Power_Exchanged = np.sum(x_list)/2
    print('Total Power Exchanged:', Total_Power_Exchanged)

    # Social Welfare value is the current value of the objective function
    print('Social Welfare:', SW)

    # print the market price
    print('Market Price:', λ)

    # calculate the total producers profit
    R1 = (λ*np.ones(len(offers)))*(x_list[:len(offers)])
    R2 = offers*offers_Power
    R = R1-R2
    Producers_Profit = 0
    for i in range (len(R)):
        if R[i] > 0:
            Producers_Profit += R[i]
    print('Producers Profit:', round(Producers_Profit, 2))

    # calculate the total consumers profit
    B1 = (λ*np.ones(len(bids)))*(x_list[len(offers):])
    B2 = bids*bids_Power
    B=[] 
    for i in range(len(B1)):
        if B1[i] != 0:
            B.append(B2[i]-B1[i])
        else: 
            B.append(0)
    Consumers_Profit = 0 
    for i in range(len(B)):
        if B[i] > 0:
            Consumers_Profit += B[i]
    print('Consumers Profit:', round(Consumers_Profit, 2))

    # calculate each producers profit
    Prod_i_profit = np.where(R<0, 0, R)
    # calculate each consumers profit
    Buyer_i_profit = B
    # calculate Type of Offer/Bid
    type_of_offer = []

    for i in range(len(offers_Power)):
        if offers_Power[i] - x_list[i] == 0:
            type_of_offer.append('Fully Accepted')
        else:
            if x_list[i] == 0:
                type_of_offer.append('Rejected')
            else:
                type_of_offer.append('Partially Accepted') 

    type_of_bid = []
    for i in range(len(bids_Power)):
        if bids_Power[i] - x_list[len(offers)+i] == 0:
            type_of_bid.append('Fully Accepted')
        else:
            if x_list[len(offers)+i] == 0:
                type_of_bid.append('Rejected')
            else:
                type_of_bid.append('Partially Accepted')
    # append results to the excel file
    output_file = 'output_files/output_%s.xlsx'%hour
    wb = load_workbook(output_file)
    ws = wb.active
    # Clear previous data in the worksheet
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            cell.value = None
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=9, max_col=11):
        for cell in row:
            cell.value = None  
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=14, max_col=ws.max_column):
        for cell in row:
            cell.value = None                
    # Write new data to the worksheet
    for i in range (len(offers)):
        ws['C%s'%(i+3)].value = x_list[i]
    for i in range (len(type_of_offer)):
        ws['D%s'%(i+3)].value = type_of_offer[i]    
    for i in range (len(offers)):
        ws['E%s'%(i+3)].value = Prod_i_profit[i]
    for i in range (len(bids)):
        ws['I%s'%(i+3)].value = x_list[len(offers)+i]    
    for i in range (len(type_of_bid)):
        ws['J%s'%(i+3)].value = type_of_bid[i] 
    for i in range (len(bids)):
        ws['K%s'%(i+3)].value = Buyer_i_profit[i]

    ws['N2'].value = Total_Power_Exchanged
    ws['N3'].value = λ
    ws['N4'].value = round(Producers_Profit, 2)
    ws['N5'].value = round(Consumers_Profit, 2)
    ws['N6'].value = SW
    ws['N8'].value = type_of_offer.count('Fully Accepted')
    ws['N9'].value = type_of_offer.count('Partially Accepted')
    ws['N10'].value = type_of_offer.count('Rejected')
    ws['N12'].value = type_of_bid.count('Fully Accepted')
    ws['N13'].value = type_of_bid.count('Partially Accepted')
    ws['N14'].value = type_of_bid.count('Rejected')

    wb.save(output_file)
    print('\n')


print('Time:', time.time()-t)

    



